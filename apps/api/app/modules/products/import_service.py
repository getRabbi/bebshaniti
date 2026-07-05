import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000

HEADER_ALIASES = {
    "product_name": "product_name",
    "পণ্যের নাম": "product_name",
    "category": "category",
    "ক্যাটাগরি": "category",
    "unit": "unit",
    "ইউনিট": "unit",
    "buying_price": "buying_price",
    "ক্রয় মূল্য": "buying_price",
    "selling_price": "selling_price",
    "বিক্রয় মূল্য": "selling_price",
    "sku": "sku",
    "barcode": "barcode",
    "brand": "brand",
    "ব্র্যান্ড": "brand",
    "opening_stock": "opening_stock",
    "ওপেনিং স্টক": "opening_stock",
    "low_stock_alert": "low_stock_alert",
    "কম স্টক সতর্কতা": "low_stock_alert",
    "wholesale_price": "wholesale_price",
    "পাইকারি মূল্য": "wholesale_price",
    "mrp": "mrp",
    "supplier": "supplier",
    "সাপ্লায়ার": "supplier",
    "rack": "rack",
    "তাক": "rack",
    "expiry_date": "expiry_date",
    "মেয়াদ": "expiry_date",
}
REQUIRED_COLUMNS = {
    "product_name",
    "category",
    "unit",
    "buying_price",
    "selling_price",
}
NUMERIC_COLUMNS = {
    "buying_price",
    "selling_price",
    "opening_stock",
    "low_stock_alert",
    "wholesale_price",
    "mrp",
}


class ImportFileError(ValueError):
    pass


def _normalized_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\ufeff", "")


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileError("CSV must be UTF-8 encoded") from exc
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportFileError("Invalid XLSX file") from exc
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True, max_row=MAX_IMPORT_ROWS + 2)]
    workbook.close()
    return rows


def _decimal(value: Any, field: str, errors: list[str]) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        errors.append(f"{field}: invalid number")
        return Decimal("0")
    if result < 0:
        errors.append(f"{field}: cannot be negative")
    return result


def _date(value: Any, errors: list[str]) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        errors.append("expiry_date: use YYYY-MM-DD")
        return None


def parse_product_file(file_name: str, content: bytes) -> list[dict[str, Any]]:
    if len(content) > MAX_IMPORT_BYTES:
        raise ImportFileError("File exceeds the 5 MB limit")
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        source_rows = _rows_from_csv(content)
    elif suffix == ".xlsx":
        source_rows = _rows_from_xlsx(content)
    else:
        raise ImportFileError("Only CSV and XLSX files are supported")
    if not source_rows:
        raise ImportFileError("The import file is empty")
    if len(source_rows) - 1 > MAX_IMPORT_ROWS:
        raise ImportFileError(f"Import is limited to {MAX_IMPORT_ROWS} rows")

    canonical_headers: list[str | None] = []
    for header in source_rows[0]:
        canonical_headers.append(HEADER_ALIASES.get(_normalized_header(header)))
    found = {header for header in canonical_headers if header}
    missing = sorted(REQUIRED_COLUMNS - found)
    if missing:
        raise ImportFileError(f"Missing required columns: {', '.join(missing)}")

    parsed: list[dict[str, Any]] = []
    for index, values in enumerate(source_rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        raw: dict[str, Any] = {}
        for column_index, header in enumerate(canonical_headers):
            if header:
                raw[header] = values[column_index] if column_index < len(values) else None
        errors: list[str] = []
        for required in ("product_name", "category", "unit"):
            if not str(raw.get(required) or "").strip():
                errors.append(f"{required}: required")
        normalized: dict[str, Any] = {
            "row_number": index,
            "product_name": str(raw.get("product_name") or "").strip(),
            "category": str(raw.get("category") or "").strip(),
            "unit": str(raw.get("unit") or "").strip(),
            "sku": str(raw.get("sku") or "").strip() or None,
            "barcode": str(raw.get("barcode") or "").strip() or None,
            "brand": str(raw.get("brand") or "").strip() or None,
            "supplier": str(raw.get("supplier") or "").strip() or None,
            "rack": str(raw.get("rack") or "").strip() or None,
        }
        for field in NUMERIC_COLUMNS:
            normalized[field] = _decimal(raw.get(field), field, errors)
        if normalized["selling_price"] == 0 and raw.get("selling_price") in (None, ""):
            errors.append("selling_price: required")
        if normalized["buying_price"] == 0 and raw.get("buying_price") in (None, ""):
            errors.append("buying_price: required")
        normalized["expiry_date"] = _date(raw.get("expiry_date"), errors)
        normalized["errors"] = errors
        parsed.append(normalized)
    return parsed


def serializable_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            key: str(value) if isinstance(value, (Decimal, date)) else value
            for key, value in row.items()
        }
        for row in rows
    ]
